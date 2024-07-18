from django.shortcuts import render,redirect
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required,user_passes_test
# Create your views here.
from main.models import CostCodeModel,ClusterModel,GSTModel,LabourModel,PartyModel,RateModel,CountModel,RCModel,EntryModel,InvoiceRCModel,InvoiceModel,ICountModel,QCountModel,QTTModel,PartyNameModel,QuotationModel
from django.http import HttpResponse,JsonResponse
from django.views.decorators.csrf import csrf_exempt

def Login(request):
    if request.method=="POST":
        username=request.POST.get('Username')
        password=request.POST.get('Password')
        user=authenticate(request,username=username,password=password)
        if user is not None:
            login(request,user)
            return redirect('/')
    return render(request,"login.html")

def Logout(request):
    logout(request)
    return redirect('/')

@login_required(login_url='Login')
def home(request):
    PC = EntryModel.objects.filter(CallType='Call',complet='0').count()
    PW = EntryModel.objects.filter(CallType='WorkOrder',complet='0').count()
    CC = EntryModel.objects.filter(complet='1').count()
    Invoice = InvoiceModel.objects.all().count()
    data={'PC':PC,'PW':PW,'CC':CC,'Invoice':Invoice}
    return render(request,'dashboard.html',data)

@login_required(login_url='Login')
def Cluster(request):
    data=ClusterModel.objects.all()
    if request.method=="POST":    
        Name=request.POST.get('Name')
        Amount=request.POST.get('Amount')
        ABH = request.POST.get('ABH')
        if ABH == 'ATM':
            ATM='True'
        else:
            ATM='False'
        if ABH ==  'Branch':
            Branch='True'
        else:
            Branch='False'
        if ABH ==  'HubLocation':
            HubLocation='True'
        else:
            HubLocation='False'
        dt=ClusterModel.objects.create(Name=Name,Amount=Amount,ATM=ATM,Branch=Branch,HubLocation=HubLocation)
        dt.save()
        return redirect('/Cluster')
    return render(request,'cluster.html',{'data':data})

@login_required(login_url='Login')
def EditCluster(request,id):
    data=ClusterModel.objects.get(id=id)
    if request.method=="POST":    
        data.Name=request.POST.get('Name')
        data.Amount=request.POST.get('Amount')
        ABH = request.POST.get('ABH')
        if ABH == 'ATM':
            data.ATM='True'
        else:
            data.ATM='False'
        if ABH ==  'Branch':
            data.Branch='True'
        else:
            data.Branch='False'
        if ABH ==  'HubLocation':
            data.HubLocation='True'
        else:
            data.HubLocation='False'
        data.save()
        return redirect('/Cluster')
    return render(request,'editcluster.html',{'data':data})

@login_required(login_url='Login')
def DeleteCluster(request,id):
    data=ClusterModel.objects.get(id=id)
    data.delete()
    return redirect('/Cluster')

@login_required(login_url='Login')
def CallEntry(request):
    data=EntryModel.objects.filter(complet='0')
    return render(request,'callentry.html',{'data':data})

@csrf_exempt
def RCCODE(request):
    if request.method == 'POST':
        RC = request.POST.get('RC')
        RCL = RateModel.objects.get(CodeNo=RC)
        return JsonResponse({'RC':RCL.Rate})

@csrf_exempt
def CCName(request):
    if request.method == 'POST':
        value = request.POST.get('value')
        RCL = CostCodeModel.objects.get(CostCode=value)
        return JsonResponse({'Name':RCL.Name})
    
@login_required(login_url='Login')
def AddCallEntry(request):
    try:
        CM=CountModel.objects.get(id=1)
        count = CM.Counter
    except CountModel.DoesNotExist:
        CM=CountModel.objects.create(Counter=1)
        count = 1
    Party=PartyModel.objects.all()
    PN={i.PartyName for i in Party}
    CC=CostCodeModel.objects.all()
    Cluster=ClusterModel.objects.all()
    RC = RateModel.objects.all()
    RCM=RCModel.objects.all()
    RCM2=RCModel.objects.filter(Counter=count)
    TAmount=0
    for i in RCM2:
        TAmount=TAmount+ eval(i.Amount)
    maindata={'Party':Party,'CC':CC,'CM':CM,'RC':RC,'RCM':RCM,'TAmount':TAmount,'PN':PN,'Cluster':Cluster}
    if request.method=="POST":
        Counter = request.POST.get('Counter')
        PartyName = request.POST.get('PartyName')
        City = request.POST.get('City')
        Branch = request.POST.get('Branch')
        Branch= Branch.replace("%20", " ")
        Code = request.POST.get('Code')
        Cluster = request.POST.get('Cluster')
        Cluster_id = request.POST.get('Cluster_id')
        Address = request.POST.get('Address')
        NO1 = request.POST.get('NO1')
        NO2 = request.POST.get('NO2')
        NO3 = request.POST.get('NO3')
        NO4 = request.POST.get('NO4')
        CallType = request.POST.get('CallType')
        Date = request.POST.get('Date')
        CallDate = request.POST.get('CallDate')
        CloseDate = request.POST.get('CloseDate')
        CodeNo = request.POST.get('CodeNo')
        CostCode = request.POST.get('CostCode')
        ContactNo = request.POST.get('ContactNo')
        ContactPerson = request.POST.get('ContactPerson')
        Email = request.POST.get('Email')
        CCName = request.POST.get('CCName')
        Problem = request.POST.get('Problem')
        ChallanNo = request.POST.get('ChallanNo')
        ChallanDate = request.POST.get('ChallanDate')
        CallAllocatedTo = request.POST.get('CallAllocatedTo')
        EstimateRecd = request.POST.get('EstimateRecd')
        WorkEngaged = request.POST.get('WorkEngaged')
        Other = request.POST.get('Other')
        Remark = request.POST.get('Remark')
        TotalAmount = request.POST.get('TotalAmount')
        if WorkEngaged != "Others":
            Other = ''
        if CallType == "WorkOrder":
            data = EntryModel(Counter=Counter,PartyName=PartyName,City=City,Branch=Branch,Code=Code,Cluster=Cluster,Cluster_id=Cluster_id,Address=Address,NO1=NO1,NO2=NO2,NO3=NO3,NO4=NO4,CallType=CallType,CallDate=CallDate,Date=Date,CloseDate=CloseDate,CodeNo=CodeNo,CostCode=CostCode,ContactNo=ContactNo,ContactPerson=ContactPerson,Email=Email,CCName=CCName,Problem=Problem,ChallanNo=ChallanNo,ChallanDate=ChallanDate,CallAllocatedTo=CallAllocatedTo,EstimateRecd=EstimateRecd,WorkEngaged=WorkEngaged,Other=Other,Remark=Remark,TotalAmount=TotalAmount)
        else:
            data = EntryModel(Counter=Counter,PartyName=PartyName,City=City,Branch=Branch,Code=Code,Cluster=Cluster,Cluster_id=Cluster_id,Address=Address,NO1=NO1,NO2=NO2,NO3=NO3,NO4=NO4,CallType=CallType,Date=Date,CloseDate=CloseDate,CodeNo=CodeNo,CostCode=CostCode,ContactNo=ContactNo,ContactPerson=ContactPerson,Email=Email,CCName=CCName,Problem=Problem,ChallanNo=ChallanNo,ChallanDate=ChallanDate,CallAllocatedTo=CallAllocatedTo,EstimateRecd=EstimateRecd,WorkEngaged=WorkEngaged,Other=Other,Remark=Remark,TotalAmount=TotalAmount)
        data.save()
        CM.Counter = int(count)+1
        CM.save()
        return redirect('/CallEntry')
    return render(request,'addcallentry.html',maindata)

@login_required(login_url='Login')
def DeleteCallEntry(request,id):
    dt = EntryModel.objects.get(id=id)
    RC = RCModel.objects.filter(Counter= dt.Counter)
    dt.delete()
    RC.delete()
    return redirect('/CallEntry')

def EditCallEntry(request,pk):
    data=EntryModel.objects.get(id=pk)
    Party=PartyModel.objects.all()
    PN={i.PartyName for i in Party}
    City = PartyModel.objects.filter(PartyName=data.PartyName)
    City={i.City for i in City}
    Branch = PartyModel.objects.filter(PartyName=data.PartyName,City=data.City)
    Branch={i.Branch for i in Branch}
    RCM2=RCModel.objects.filter(Counter=data.Counter)
    CC=CostCodeModel.objects.all()
    Cluster=ClusterModel.objects.all()
    RC = RateModel.objects.all()
    RCM=RCModel.objects.all()
    TAmount=0
    for i in RCM2:
        TAmount=TAmount+ eval(i.Amount)
    if request.method=="POST":
        data=EntryModel.objects.get(id=pk)
        data.Counter = request.POST.get('Counter')
        data.PartyName = request.POST.get('PartyName')
        data.City = request.POST.get('City')
        Branch = request.POST.get('Branch')
        data.Branch= Branch.replace("%20", " ")
        data.Code = request.POST.get('Code')
        data.Cluster = request.POST.get('Cluster')
        data.Cluster_id = request.POST.get('Cluster_id')
        data.Address = request.POST.get('Address')
        data.NO1 = request.POST.get('NO1')
        data.NO2 = request.POST.get('NO2')
        data.NO3 = request.POST.get('NO3')
        data.NO4 = request.POST.get('NO4')
        CallType = request.POST.get('CallType')
        data.CallType = CallType
        if CallType == 'WorkOrder':
            data.CallDate = request.POST.get('CallDate')
        Date = request.POST.get('Date')
        data.Date = Date
        data.CloseDate = request.POST.get('CloseDate')
        data.CodeNo = request.POST.get('CodeNo')
        data.CostCode = request.POST.get('CostCode')
        data.ContactNo = request.POST.get('ContactNo')
        data.ContactPerson = request.POST.get('ContactPerson')
        data.Email = request.POST.get('Email')
        data.CCName = request.POST.get('CCName')
        data.Problem = request.POST.get('Problem')
        data.ChallanNo = request.POST.get('ChallanNo')
        data.ChallanDate = request.POST.get('ChallanDate')
        data.CallAllocatedTo = request.POST.get('CallAllocatedTo')
        data.EstimateRecd = request.POST.get('EstimateRecd')
        WorkEngaged = request.POST.get('WorkEngaged')
        data.WorkEngaged = WorkEngaged
        if WorkEngaged != "Others":
            data.Other = ''
        else:
            data.Other = request.POST.get('Other')
        data.Remark = request.POST.get('Remark')
        data.TotalAmount = request.POST.get('TotalAmount')
        data.save()
        return redirect('/CallEntry')
    return render(request,'editcallentry.html',{'data':data,'RCM':RCM,'PN':PN,'City':City,'Branch':Branch,'Cluster':Cluster,'RC':RC,'CC':CC,'TAmount':TAmount})

@csrf_exempt
def Citys(request):
    PN = request.POST.get('value')
    City = PartyModel.objects.filter(PartyName=PN)
    PN={i.City for i in City}
    City = list(PN)
    return JsonResponse({'City':City})

def Branch(request,City):
    Branch = PartyModel.objects.filter(City=City)
    PN={i.Branch for i in Branch}
    Branch = list(PN)
    return JsonResponse({'Branch':Branch})

def RCCreat(request):
    if request.method == 'POST':
        sid = request.POST.get('sid')
        Counter = request.POST['Counter']
        RCCode = request.POST['RCCode']
        RCDescription = request.POST['RCDescription']
        Quantity = request.POST['Quantity']
        Rate = request.POST['Rate']
        Labour = request.POST['Labour']
        Amount = request.POST['Amount']
        RC = RateModel.objects.get(CodeNo=RCCode)
        if sid == '':
            data=RCModel(Counter=Counter,RCCode=RCCode,Unit=RC.Unit,HSNCode=RC.HSNCode,RCDescription=RCDescription,Quantity=Quantity,Rate=Rate,Labour=Labour,Amount=Amount)
        else:
            data=RCModel(id=sid,Counter=Counter,RCCode=RCCode,Unit=RC.Unit,HSNCode=RC.HSNCode,RCDescription=RCDescription,Quantity=Quantity,Rate=Rate,Labour=Labour,Amount=Amount)
        data.save()
        # data load
        show=RCModel.objects.filter(Counter=Counter).values()
        RCM2=RCModel.objects.filter(Counter=Counter)
        TAmount=0
        for i in RCM2:
            TAmount=TAmount+ eval(i.Amount)
        show=list(show)
        return JsonResponse({'show':show,'TAmount':TAmount})

def RCEdit(request):
    if request.method == 'POST':
        id = request.POST.get('sid')
        data=RCModel.objects.get(pk=id)
        mydata={"id":data.id,"Counter":data.Counter,"RCCode":data.RCCode,"RCDescription":data.RCDescription,"Quantity":data.Quantity,"Rate":data.Rate,"Labour":data.Labour,"Amount":data.Amount}
        return JsonResponse(mydata)

def RCDelete(request):
    if request.method == 'POST':
        id = request.POST.get('sid')
        data=RCModel.objects.get(pk=id)
        i=data.Counter
        data.delete()
        RCM2=RCModel.objects.filter(Counter=i)
        TAmount=0
        for i in RCM2:
            TAmount=TAmount+ eval(i.Amount)
        return JsonResponse({'status':1,'TAmount':TAmount})
    else:
        return JsonResponse({'status':0})

def EntryClase(request):
    CM=CountModel.objects.get(id=1)
    RCM=RCModel.objects.filter(Counter=CM.Counter)
    RCM.delete()
    return redirect('/CallEntry')

@login_required(login_url='Login')
def PendingCall(request):
    data=EntryModel.objects.filter(CallType='Call',complet='0')
    return render(request,'pendingcall.html',{'data':data})

@login_required(login_url='Login')
def PendingWorkOrder(request):
    data=EntryModel.objects.filter(CallType='WorkOrder',complet='0')
    return render(request,'pendingworkorder.html',{'data':data})

@login_required(login_url='Login')
def CompleteCall(request):
    data=EntryModel.objects.filter(complet='1')
    return render(request,'completecall.html',{'data':data})

@login_required(login_url='Login')
def CostCode(request):
    data=CostCodeModel.objects.all()
    if request.method=="POST":    
        CostCode=request.POST.get('CostCode')
        Name=request.POST.get('Name')
        Remarks=request.POST.get('Remarks')
        dt=CostCodeModel.objects.create(CostCode=CostCode,Name=Name,Remarks=Remarks)
        dt.save()
        return redirect('/CostCode')
    return render(request,'costcode.html',{'data':data})

@login_required(login_url='Login')
def EditCostCode(request,id):
    data=CostCodeModel.objects.get(id=id)
    if request.method=="POST":    
        data.CostCode=request.POST.get('CostCode')
        data.Name=request.POST.get('Name')
        data.Remarks=request.POST.get('Remarks')
        data.save()
        return redirect('/CostCode')
    return render(request,'editcostcode.html',{'data':data})

@login_required(login_url='Login')
def DeleteCostCode(request,id):
    data=CostCodeModel.objects.get(id=id)
    data.delete()
    return redirect('/CostCode')

@login_required(login_url='Login')
def GST(request):
    data=GSTModel.objects.all()
    if request.method=="POST":    
        GSTName=request.POST.get('GSTName')
        SGST=request.POST.get('SGST')
        CGST=request.POST.get('CGST')
        dt=GSTModel.objects.create(GSTName=GSTName,SGST=SGST,CGST=CGST)
        dt.save()
        return redirect('/GST')
    return render(request,'GST.html',{'data':data})

@login_required(login_url='Login')
def EditGST(request,id):
    data=GSTModel.objects.get(id=id)
    if request.method=="POST":    
        data.GSTName=request.POST.get('GSTName')
        data.SGST=request.POST.get('SGST')
        data.CGST=request.POST.get('CGST')
        data.save()
        return redirect('/GST')
    return render(request,'editgst.html',{'data':data})

@login_required(login_url='Login')
def DeleteGST(request,id):
    data=GSTModel.objects.get(id=id)
    data.delete()
    return redirect('/GST')

@login_required(login_url='Login')
def Party(request):
    data=PartyModel.objects.all()
    return render(request,'party.html',{'data':data})

@login_required(login_url='Login')
def AddParty(request):
    data=ClusterModel.objects.all()
    PN=PartyNameModel.objects.all()
    if request.method=="POST":    
        Code=request.POST.get('Code')
        PartyName=request.POST.get('PartyName')
        id=request.POST.get('Cluster')
        cl= ClusterModel.objects.get(id=id)
        Cluster = cl.Name
        Active = id
        City=request.POST.get('City')
        Branch=request.POST.get('Branch')
        Address=request.POST.get('Address')
        GSTNo=request.POST.get('GSTNo')
        NO1=request.POST.get('NO1')
        NO2=request.POST.get('NO2')
        NO3=request.POST.get('NO3')
        NO4=request.POST.get('NO4')
        NO5=request.POST.get('NO5')
        NO6=request.POST.get('NO6')
        dt=PartyModel.objects.create(Code=Code,PartyName=PartyName,Cluster=Cluster,Active=Active,City=City,Branch=Branch,Address=Address,GSTNo=GSTNo,NO1=NO1,NO2=NO2,NO3=NO3,NO4=NO4,NO5=NO5,NO6=NO6)
        dt.save()
        return redirect('/Party')
    return render(request,'addparty.html',{'data':data,'PN':PN})

@login_required(login_url='Login')
def EditParty(request,id):
    data=PartyModel.objects.get(id=id)
    Clr=ClusterModel.objects.all()
    if request.method=="POST":    
        data.Code=request.POST.get('Code')
        data.PartyName=request.POST.get('PartyName')
        # data.Cluster=request.POST.get('Cluster')
        id=request.POST.get('Cluster')
        cl= ClusterModel.objects.get(id=id)
        data.Cluster = cl.Name
        data.Active = id
        data.City=request.POST.get('City')
        data.Branch=request.POST.get('Branch')
        data.Address=request.POST.get('Address')
        data.GSTNo=request.POST.get('GSTNo')
        data.NO1=request.POST.get('NO1')
        data.NO2=request.POST.get('NO2')
        data.NO3=request.POST.get('NO3')
        data.NO4=request.POST.get('NO4')
        data.NO5=request.POST.get('NO5')
        data.NO6=request.POST.get('NO6')
        data.save()
        return redirect('/Party')
    return render(request,'editparty.html',{'data':data,'Clr':Clr})

@login_required(login_url='Login')
def DeleteParty(request,id):
    data=PartyModel.objects.get(id=id)
    data.delete()
    return redirect('/Party')

@login_required(login_url='Login')
def Labour(request):
    data=LabourModel.objects.all()
    return render(request,'labour.html',{'data':data})

@login_required(login_url='Login')
def AddLabour(request):
    if request.method=="POST":    
        Name=request.POST.get('Name')
        Address=request.POST.get('Address')
        Mobile1=request.POST.get('MobileNo1')
        Mobile2=request.POST.get('MobileNo1')
        Remarks=request.POST.get('Remarks')
        dt=LabourModel.objects.create(Name=Name,Address=Address,Mobile1=Mobile1,Mobile2=Mobile2,Remarks=Remarks)
        dt.save()
        return redirect('/Labour')
    return render(request,'addlabour.html')

@login_required(login_url='Login')
def EditLabour(request,id):
    data=LabourModel.objects.get(id=id)
    if request.method=="POST":    
        data.Name=request.POST.get('Name')
        data.Address=request.POST.get('Address')
        data.Mobile1=request.POST.get('MobileNo1')
        data.Mobile2=request.POST.get('MobileNo2')
        data.Remarks=request.POST.get('Remarks')
        data.save()
        return redirect('/Labour')
    return render(request,'editlabour.html',{'data':data})

@login_required(login_url='Login')
def DeleteLabour(request,id):
    data=LabourModel.objects.get(id=id)
    data.delete()
    return redirect('/Labour')

@login_required(login_url='Login')
def Rate(request):
    data=RateModel.objects.all()
    return render(request,'rate.html',{'data':data})

@login_required(login_url='Login')
def AddRate(request):
    data=GSTModel.objects.all()
    if request.method=="POST":    
        CodeNo=request.POST.get('CodeNo')
        Description=request.POST.get('Description')
        HSNCode=request.POST.get('HSNCode')
        Unit=request.POST.get('Unit')
        Rate=request.POST.get('Rate')
        Remarks=request.POST.get('Remarks')
        GSTName=request.POST.get('GSTName')
        data=GSTModel.objects.get(GSTName=GSTName)
        SGST=data.SGST
        CGST=data.CGST    
        IGSTName=request.POST.get('IGSTName')
        IGST=0
        dt=RateModel.objects.create(CodeNo=CodeNo,Description=Description,HSNCode=HSNCode,Unit=Unit,Rate=Rate,Remarks=Remarks,GSTName=GSTName,SGST=SGST,CGST=CGST,IGSTName=IGSTName,IGST=IGST)
        dt.save()
        return redirect('/RCCode')
    return render(request,'addrate.html',{'data':data})

@login_required(login_url='Login')
def EditRate(request,id):
    gst=GSTModel.objects.all()
    data=RateModel.objects.get(id=id)
    if request.method=="POST":    
        data.CodeNo=request.POST.get('CodeNo')
        data.Description=request.POST.get('Description')
        data.HSNCode=request.POST.get('HSNCode')
        data.Unit=request.POST.get('Unit')
        data.Rate=request.POST.get('Rate')
        data.Remarks=request.POST.get('Remarks')
        GSTName=request.POST.get('GSTName')
        data.GSTName=GSTName
        dt=GSTModel.objects.get(GSTName=GSTName)
        data.SGST=dt.SGST
        data.CGST=dt.CGST    
        data.IGSTName=request.POST.get('IGSTName')
        data.IGST=0
        data.save()
        return redirect('/RCCode')
    return render(request,'editrate.html',{'data':data,'gst':gst})

@login_required(login_url='Login')
def DeleteRate(request,id):
    data=RateModel.objects.get(id=id)
    data.delete()
    return redirect('/RCCode')

@login_required(login_url='Login')
def Invoice(request):
    dt=InvoiceModel.objects.all()
    data={'dt':dt}
    return render(request,'invoice.html',data)

@login_required(login_url='Login')
def AddInvoice(request):
    try:
        CM=ICountModel.objects.get(id=1)
        InvoiceNo = CM.InvoiceNo
    except ICountModel.DoesNotExist:
        CM=ICountModel.objects.create(InvoiceNo=1)
        InvoiceNo = 1
    PN=PartyModel.objects.all()
    PN={i.PartyName for i in PN} 
    maindata={"PN":PN,'InvoiceNo':InvoiceNo}
    if request.method=="POST":    
        PartyName=request.POST.get('PartyName')
        datein=request.POST.get('datein')
        Type=request.POST.get('Type')
        if datein == 'Date':
            FromDate=request.POST.get('FromDate')
            ToDate=request.POST.get('ToDate')
            show=EntryModel.objects.filter(PartyName=PartyName,CallType=Type,Date__range=[FromDate, ToDate],complet='0')
            ttl = 0
            for i in show:
                ttl += eval(i.TotalAmount)
            ttl = "%.2f" % ttl
            maindata={"PN":PN,'InvoiceNo':InvoiceNo,'show':show,'PartyName':PartyName,'Type':Type,'datein':datein,'FromDate':FromDate,'ToDate':ToDate,'ttl':ttl}
        else:
            show=EntryModel.objects.filter(PartyName=PartyName,CallType=Type,complet='0')
            ttl = 0
            for i in show:
                ttl += eval(i.TotalAmount)
            ttl = "%.2f" % ttl
            maindata={"PN":PN,'InvoiceNo':InvoiceNo,'show':show,'PartyName':PartyName,'Type':Type,'datein':datein,'ttl':ttl}
    return render(request,'addinvoice.html',maindata)

def PartyName(request,PN):
    show=EntryModel.objects.filter(PartyName=PN).values()
    show=list(show)
    return JsonResponse({'show':show})

@login_required(login_url='Login')
def AddInvoiceMaIN(request):
    if request.method=="POST":
        InvoiceData=request.POST.get('InvoiceData')
        InvoiceNo=request.POST.get('InvoiceNo')
        BillMonth=request.POST.get('BillMonth')
        BillYear=request.POST.get('BillYear')
        Tax=request.POST.get('Tax')
        TotalA=request.POST.get('TotalAmount')
        PartyName=request.POST.get('PartyName')
        datein=request.POST.get('datein')
        Type=request.POST.get('Type')
        if TotalA:
            TA = eval(TotalA)
        else:
            TA = 0
        if 0 < TA:
            if datein == 'Date':
                FromDate=request.POST.get('FromDate')
                ToDate=request.POST.get('ToDate')
                show=EntryModel.objects.filter(PartyName=PartyName,CallType=Type,Date__range=[FromDate, ToDate],complet='0')
                for i in show:
                    dt = RCModel.objects.filter(Counter=i.Counter)
                    GA=0
                    for j in dt:
                        Rate = RateModel.objects.get(CodeNo=j.RCCode)
                        GSTRate = Rate.SGST
                        GSTAmount = eval(j.Amount)* int(GSTRate) / 100
                        GSTTA = GSTAmount + GSTAmount
                        GA +=GSTTA
                        TotalAmount = eval(j.Amount) + GSTTA
                        GSTTA = "%.2f" % GSTTA
                        GSTAmount = "%.2f" % GSTAmount
                        TotalAmount = "%.2f" % TotalAmount
                        IRCM = InvoiceRCModel.objects.create(InvoiceNo=InvoiceNo,Counter=j.Counter,RCCode=j.RCCode,RCDescription=j.RCDescription,Quantity=j.Quantity,Rate=j.Rate,Labour=j.Labour,Amount=j.Amount,GSTRate=GSTRate,GSTAmount=GSTAmount,GSTTA=GSTTA,TotalAmount=TotalAmount)
                        IRCM.save()
                    GA = GA
                    TotalAmount = GA + eval(TotalA)
                    TotalAmount = "%.2f" % TotalAmount
                    i.complet = '1'
                    i.save()
                MainInvoice = InvoiceModel.objects.create(PartyName=PartyName,InvoiceData=InvoiceData,InvoiceNo=InvoiceNo,BillMonth=BillMonth,BillYear=BillYear,Tax=Tax,datein=datein,Type=Type,FromDate=FromDate,ToDate=ToDate,Amount=TotalA,GSTAmount=GA,TotalAmount=TotalAmount)
                MainInvoice.save()
                CM = ICountModel.objects.get(id=1)
                CM.InvoiceNo = int(InvoiceNo)+1
                CM.save()
                return redirect("/AddInvoice")
            else:
                show=EntryModel.objects.filter(PartyName=PartyName,CallType=Type,complet='0')
                for i in show:
                    dt = RCModel.objects.filter(Counter=i.Counter)
                    GA=0
                    for j in dt:
                        Rate = RateModel.objects.get(CodeNo=j.RCCode)
                        GSTRate = Rate.SGST
                        GSTAmount = eval(j.Amount) * int(GSTRate) / 100
                        GSTTA = GSTAmount + GSTAmount
                        GA +=GSTTA
                        TotalAmount = eval(j.Amount) + GSTTA
                        GSTTA = "%.2f" % GSTTA
                        GSTAmount = "%.2f" % GSTAmount
                        TotalAmount = "%.2f" % TotalAmount
                        IRCM = InvoiceRCModel.objects.create(InvoiceNo=InvoiceNo,Counter=j.Counter,RCCode=j.RCCode,Unit=j.Unit,HSNCode=j.HSNCode,RCDescription=j.RCDescription,Quantity=j.Quantity,Rate=j.Rate,Labour=j.Labour,Amount=j.Amount,GSTRate=GSTRate,GSTAmount=GSTAmount,GSTTA=GSTTA,TotalAmount=TotalAmount)
                        IRCM.save()
                    GA = GA
                    TotalAmount = GA + eval(TotalA)
                    TotalAmount = "%.2f" % TotalAmount
                    i.complet = '1'
                    i.save()
                MainInvoice = InvoiceModel.objects.create(PartyName=PartyName,InvoiceData=InvoiceData,InvoiceNo=InvoiceNo,BillMonth=BillMonth,BillYear=BillYear,Tax=Tax,datein=datein,Type=Type,Amount=TotalA,GSTAmount=GA,TotalAmount=TotalAmount)
                MainInvoice.save()
                CM = ICountModel.objects.get(id=1)
                CM.InvoiceNo = int(InvoiceNo)+1
                CM.save()
                return redirect("/AddInvoice")
        else:
            return redirect("/AddInvoice")
        
@login_required(login_url='Login')
def QuotationEntry(request):
    data=QuotationModel.objects.filter()
    return render(request,'quotation.html',{'data':data})

@login_required(login_url='Login')
def AddQuotation(request):
    try:
        CM=QCountModel.objects.get(id=1)
        QuotationNo = CM.QuotationNo
    except QCountModel.DoesNotExist:
        CM=QCountModel.objects.create(QuotationNo=1)
        QuotationNo = 1
    PN=PartyModel.objects.all()
    PN={i.PartyName for i in PN} 
    CC=CostCodeModel.objects.all()
    Cluster=ClusterModel.objects.all()
    RC = RateModel.objects.all()
    RCM=QTTModel.objects.all()
    RCM2=QTTModel.objects.filter(Counter=QuotationNo)
    TAmount=0
    for i in RCM2:
        TAmount=TAmount+ eval(i.Amount)
    maindata={'CC':CC,'CM':CM,'RC':RC,'RCM':RCM,'RCM2':RCM2,'QuotationNo':QuotationNo,'TAmount':TAmount,'PN':PN,'Cluster':Cluster}
    if request.method=="POST":
        Counter = request.POST.get('Counter')
        QuotationNo = request.POST.get('QuotationNo')
        Date = request.POST.get('Date')
        PartyName = request.POST.get('PartyName')
        City = request.POST.get('City')
        Branch = request.POST.get('Branch')
        Branch= Branch.replace("%20", " ")
        Code = request.POST.get('Code')
        Address = request.POST.get('Address')
        Problem = request.POST.get('Problem')
        CallRef = request.POST.get('CallRef')
        CallNo = request.POST.get('CallNo')
        WorkEngaged = request.POST.get('WorkEngaged')
        Other = request.POST.get('Other')
        Subject = request.POST.get('Subject')
        AreaOfWork = request.POST.get('AreaOfWork')
        Remark = request.POST.get('Remark')
        TotalAmount = request.POST.get('TotalAmount')
        TC = request.POST.get('TC')
        if WorkEngaged != "Others":
            Other = ''
        dt=QuotationModel.objects.create(Counter=Counter,QuotationNo=QuotationNo,Date=Date,PartyName=PartyName,City=City,Branch=Branch,Address=Address,Problem=Problem,Code=Code,CallRef=CallRef,CallNo=CallNo,WorkEngaged=WorkEngaged,Other=Other,Subject=Subject,AreaOfWork=AreaOfWork,Remark=Remark,TC=TC,TotalAmount=TotalAmount)
        dt.save()
        CM.QuotationNo = int(QuotationNo)+1
        CM.save()
        return redirect('/AddQuotation')
    return render(request,'addquotation.html',maindata)

@login_required(login_url='Login')
def DeleteQuotation(request,id):
    data=QuotationModel.objects.get(id=id)
    RC=QTTModel.objects.filter(Counter=data.Counter)
    RC.delete()
    data.delete()
    return redirect('/Quotation')

@login_required(login_url='Login')
def EditQuotation(request,id):
    dt = QuotationModel.objects.get(id=id)
    PN=PartyModel.objects.all()
    PN={i.PartyName for i in PN}
    City = PartyModel.objects.filter(PartyName=dt.PartyName)
    City={i.City for i in City}
    Branch = PartyModel.objects.filter(PartyName=dt.PartyName,City=dt.City)
    Branch={i.Branch for i in Branch}
    CC=CostCodeModel.objects.all()
    Cluster=ClusterModel.objects.all()
    RC = RateModel.objects.all()
    RCM=QTTModel.objects.all()
    RCM2=QTTModel.objects.filter(Counter=dt.QuotationNo)
    TAmount=0
    for i in RCM2:
        TAmount=TAmount+ eval(i.Amount)
    maindata={'CC':CC,'dt':dt,'RC':RC,'RCM':RCM,'RCM2':RCM2,'Branch':Branch,'City':City,'QuotationNo':dt.QuotationNo,'TAmount':TAmount,'PN':PN,'Cluster':Cluster}
    if request.method=="POST":
        dt = QuotationModel.objects.get(id=id)
        dt.Counter = request.POST.get('Counter')
        dt.QuotationNo = request.POST.get('QuotationNo')
        dt.Date = request.POST.get('Date')
        dt.PartyName = request.POST.get('PartyName')
        dt.City = request.POST.get('City')
        Branch = request.POST.get('Branch')
        dt.Branch= Branch.replace("%20", " ")
        dt.Code = request.POST.get('Code')
        dt.Address = request.POST.get('Address')
        dt.Problem = request.POST.get('Problem')
        dt.CallRef = request.POST.get('CallRef')
        dt.CallNo = request.POST.get('CallNo')
        WorkEngaged = request.POST.get('WorkEngaged')
        dt.WorkEngaged = WorkEngaged
        if WorkEngaged != "Others":
            dt.Other = ''
        else:
            dt.Other = request.POST.get('Other')
        dt.Subject = request.POST.get('Subject')
        dt.AreaOfWork = request.POST.get('AreaOfWork')
        dt.Remark = request.POST.get('Remark')
        dt.TotalAmount = request.POST.get('TotalAmount')
        dt.TC = request.POST.get('TC')
        dt.save()
        return redirect('/Quotation')
    return render(request,'editquotation.html',maindata)

def QRCCreat(request):
    if request.method == 'POST':
        sid = request.POST.get('sid')
        Counter = request.POST['Counter']
        TYPES = request.POST['TYPES']
        RCCode = request.POST['RCCode']
        RCDescription = request.POST['RCDescription']
        Quantity = request.POST['Quantity']
        Rate = request.POST['Rate']
        Labour = request.POST['Labour']
        Amount = request.POST['Amount']
        if sid == '':
            data=QTTModel(Counter=Counter,RCCode=RCCode,RCDescription=RCDescription,Quantity=Quantity,Rate=Rate,Labour=Labour,Amount=Amount,TYPES=TYPES)
        else:
            data=QTTModel(id=sid,Counter=Counter,RCCode=RCCode,RCDescription=RCDescription,Quantity=Quantity,Rate=Rate,Labour=Labour,Amount=Amount,TYPES=TYPES)
        data.save()
        # data load
        show=QTTModel.objects.filter(Counter=Counter).values()
        RCM2=QTTModel.objects.filter(Counter=Counter)
        TAmount=0
        for i in RCM2:
            TAmount=TAmount+ eval(i.Amount)
        show=list(show)
        return JsonResponse({'show':show,'TAmount':TAmount})
    
def QRCEdit(request):
    if request.method == 'POST':
        id = request.POST.get('sid')
        data=QTTModel.objects.get(pk=id)
        mydata={"id":data.id,"Counter":data.Counter,"RCCode":data.RCCode,"RCDescription":data.RCDescription,"Quantity":data.Quantity,"Rate":data.Rate,"Labour":data.Labour,"Amount":data.Amount,'TYPES':data.TYPES}
        return JsonResponse(mydata)

def QRCDelete(request):
    if request.method == 'POST':
        id = request.POST.get('sid')
        data=QTTModel.objects.get(pk=id)
        i=data.Counter
        data.delete()
        RCM2=QTTModel.objects.filter(Counter=i)
        TAmount=0
        for i in RCM2:
            TAmount=TAmount+ eval(i.Amount)
        return JsonResponse({'status':1,'TAmount':TAmount})
    else:
        return JsonResponse({'status':0})
    
@login_required(login_url='Login')
def ClaseQuotation(request,id):
        data=QTTModel.objects.filter(Counter=id)
        data.delete()
        return redirect('/Quotation')

@login_required(login_url='Login')
def TaxInvoice(request,id):
    dt=InvoiceModel.objects.get(id=id)
    party = PartyModel.objects.all()
    Rc = InvoiceRCModel.objects.filter(InvoiceNo=dt.InvoiceNo)
    Amount = 0
    GSTAmount = 0
    TotalAmount = 0
    for j in Rc:
        Amount += eval(j.Amount)
        GSTAmount += eval(j.GSTAmount)
        TotalAmount += eval(j.TotalAmount)
    Amount = "%.2f" % Amount
    GSTAmount = "%.2f" % GSTAmount
    TotalAmount = "%.2f" % TotalAmount
    p = {}
    for i in party:
        p['PartyName'] = i.PartyName
        break
    main={'dt':dt,'party':party,'p':p,'Rc':Rc,'Amount':Amount,'GSTAmount':GSTAmount,'TotalAmount':TotalAmount}
    return render(request,'taxinvoice.html',main)

@csrf_exempt
def CodeNoC(request):
    if request.method == 'POST':
        CN = request.POST.get('CN')
        RCL = EntryModel.objects.filter(CodeNo=CN).exists()
        if RCL == True:
            RC = 1
        else:
            RC = 0
        return JsonResponse({'RC':RC})
    
@csrf_exempt
def PartyData(request):
    if request.method == 'POST':
        Party = request.POST.get('Party')
        City = request.POST.get('City')
        Branch = request.POST.get('Branch')
        Branch2= Branch.replace("%20", " ")
        RC = PartyModel.objects.get(PartyName=Party,City=City,Branch=Branch2)
        CodeNo=RC.Code
        Address=RC.Address
        Cluster=RC.Cluster
        AID=RC.Active
        cl= ClusterModel.objects.get(id=RC.Active)
        if cl.ATM == 'True':
            Active='ATM'
        elif cl.Branch == 'True':
            Active='Branch'
        elif cl.HubLocation == 'True':
            Active='Hub Location'
        return JsonResponse({'RC':CodeNo,'AD':Address,'CL':Cluster,'AL':Active,'AID':AID})

@login_required(login_url='Login')
def PartyName(request):
    data=PartyNameModel.objects.all()
    if request.method=="POST":    
        PartyName=request.POST.get('PartyName')
        dt=PartyNameModel.objects.create(Name=PartyName)
        dt.save()
        return redirect('/PartyName')
    return render(request,'partyname.html',{'data':data})

@login_required(login_url='Login')
def EditPartyName(request,id):
    data=PartyNameModel.objects.get(id=id)
    if request.method=="POST":    
        data.Name=request.POST.get('PartyName')
        data.save()
        return redirect('/PartyName')
    return render(request,'editpartyname.html',{'data':data})

@login_required(login_url='Login')
def DeletePartyName(request,id):
    data=PartyNameModel.objects.get(id=id)
    data.delete()
    return redirect('/PartyName')

@login_required(login_url='Login')
def Annexure(request):
    CM=ClusterModel.objects.all()
    maindata={"CM":CM}
    if request.method=="POST":    
        ClusterID=request.POST.get('ClusterName')
        if ClusterID == '--Select--':
            return redirect('/ClusterReport')
        datein=request.POST.get('datein')
        Type=request.POST.get('Type')
        if datein == 'Date':
            FromDate=request.POST.get('FromDate')
            ToDate=request.POST.get('ToDate')
            show=EntryModel.objects.filter(CallType=Type,Date__range=[FromDate, ToDate],complet='1',Cluster_id=ClusterID)
            maindata={'show':show,'Type':Type,'datein':datein,'FromDate':FromDate,'ToDate':ToDate,"CM":CM,'ClusterID':int(ClusterID)}
        else:
            show=EntryModel.objects.filter(CallType=Type,complet='1',Cluster_id=ClusterID)
            maindata={'show':show,'Type':Type,'datein':datein,"CM":CM,'ClusterID':int(ClusterID)}
    return render(request,'clusterreport.html',maindata)
    # return render(request,'partyname.html',{'data':data})

def ClusterEL(request):
    if request.method=="POST":    
        ClusterID=request.POST.get('ClusterName')
    if ClusterID == '--Select--':
        return redirect('/ClusterReport')
    datein=request.POST.get('datein')
    Type=request.POST.get('Type')
    if datein == 'Date':
        FromDate=request.POST.get('FromDate')
        ToDate=request.POST.get('ToDate')
        show=EntryModel.objects.filter(CallType=Type,Date__range=[FromDate, ToDate],complet='1',Cluster_id=ClusterID)
        CM=ClusterModel.objects.get(id=ClusterID)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Cluster-Report-'+CM.Name+'-'+CM.Type+'.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = f"Cluster-Report"
        ws.merge_cells('A1:T1')
        ws['A1'] = 'Name of Agency/Vendor  :-  Giri Associates'
        ws['A2'] = 'Name Of Cluster & Type :- '+ CM.Name +' , '+ CM.Type
         # Add headers
        headers = ['Sr.No.',"Description of Goods","Work start- date","Work end-date","Challan No.","Challan date","call log.no.","call log. Date","RC Code","HSN no.","Location Code","UOM",'Qty.','Rate','Amount',"SGST Rate","SGST Amount",'CGST Rate','CGST Amount','Total Amount payable']
        ws.append(headers)
        for cell in ws["3:3"]:
            cell.font = Font(bold=True)
        num=1
        for i in show:
            dt=InvoiceRCModel.objects.filter(Counter=i.Counter)
            for j in dt:
                ws.append([num,j.RCDescription,i.Date.strftime('%d-%m-%Y'),i.CloseDate.strftime('%d-%m-%Y'),i.ChallanNo,i.ChallanDate.strftime('%d-%m-%Y'),i.CodeNo,i.Date.strftime('%d-%m-%Y'),j.RCCode,j.HSNCode,num+1000,j.Unit,j.Quantity,j.Rate,j.Amount,j.GSTRate + " %",j.GSTAmount,j.GSTRate + " %",j.GSTAmount,j.TotalAmount])
                num+=1
            Amount = 0
            GSTAmount = 0
            TotalAmount = 0
            for j in dt:
                Amount += eval(j.Amount)
                GSTAmount += eval(j.GSTAmount)
                TotalAmount += eval(j.TotalAmount)
            Amount = "%.2f" % Amount
            GSTAmount = "%.2f" % GSTAmount
            TotalAmount = "%.2f" % TotalAmount
            ws.append(['','','','','','','','','','','','','','',Amount,'',GSTAmount,'',GSTAmount,TotalAmount ])
            ws.append(['','','','','','','','','','','','','','','','','','','','' ])
        side = Side(border_style='thin', color='000000')
        set_border(ws, side)
        # Save the workbook to the HttpResponse
        wb.save(response)
        return response
    else:
        show=EntryModel.objects.filter(CallType=Type,complet='1',Cluster_id=ClusterID)
        CM=ClusterModel.objects.get(id=ClusterID)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Cluster-Report-'+CM.Name+'-'+CM.Type+'.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = f"Cluster-Report"
        ws.merge_cells('A1:T1')
        ws['A1'] = 'Name Of Agency/Vendor  :-  Giri Associates'
        ws['A2'] = 'Name Of Cluster & Type :- '+ CM.Name +' , '+ CM.Type
         # Add headers
        headers = ['Sr.No.',"Description of Goods","Work start- date","Work end-date","Challan No.","Challan date","call log.no.","call log. Date","RC Code","HSN no.","Location Code","UOM",'Qty.','Rate','Amount',"SGST Rate","SGST Amount",'CGST Rate','CGST Amount','Total Amount payable']
        ws.append(headers)
        for cell in ws["3:3"]:
            cell.font = Font(bold=True)
        num=1
        for i in show:
            dt=InvoiceRCModel.objects.filter(Counter=i.Counter)
            for j in dt:
                ws.append([num,j.RCDescription,i.Date.strftime('%d-%m-%Y'),i.CloseDate.strftime('%d-%m-%Y'),i.ChallanNo,i.ChallanDate.strftime('%d-%m-%Y'),i.CodeNo,i.Date.strftime('%d-%m-%Y'),j.RCCode,j.HSNCode,num+1000,j.Unit,j.Quantity,j.Rate,j.Amount,j.GSTRate + " %",j.GSTAmount,j.GSTRate + " %",j.GSTAmount,j.TotalAmount])
                num+=1
            Amount = 0
            GSTAmount = 0
            TotalAmount = 0
            for j in dt:
                Amount += eval(j.Amount)
                GSTAmount += eval(j.GSTAmount)
                TotalAmount += eval(j.TotalAmount)
            Amount = "%.2f" % Amount
            GSTAmount = "%.2f" % GSTAmount
            TotalAmount = "%.2f" % TotalAmount
            ws.append(['','','','','','','','','','','','','','',Amount,'',GSTAmount,'',GSTAmount,TotalAmount ])
            ws.append(['','','','','','','','','','','','','','','','','','','','' ])
        side = Side(border_style='thin', color='000000')
        set_border(ws, side)
        # Save the workbook to the HttpResponse
        wb.save(response)
        return response
    

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Side, Border,Font

def set_border(ws, side=None, blank=True):
    wb = ws._parent
    side = side if side else Side(border_style='thin', color='000000')
    for cell in ws._cells.values():
        cell.border = Border(top=side, bottom=side, left=side, right=side)
    if blank:
        white = Side(border_style='thin', color='FFFFFF')
        wb._borders.append(Border(top=white, bottom=white, left=white, right=white))
        wb._cell_styles[0].borderId = len(wb._borders) - 1

def export_to_excel(request,id):
    dt=InvoiceModel.objects.get(id=id)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Annexure-{dt.Type}-{dt.InvoiceNo}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Annexure_{id}"
    ws.merge_cells('A1:T1')
    ws['A1'] = 'Name of Agency/Vendor  :-  Giri Associates'
    dt=InvoiceModel.objects.get(id=id)
    ws.merge_cells('A2:T2')
    ws['A2'] = f'Bill Month :- {dt.BillMonth} {dt.BillYear}'
    ws.merge_cells('A3:T3')
    ws['A3'] = f'Bill No : GA20242{int(dt.InvoiceNo)+5000} Dated :'+ dt.InvoiceData.strftime('%d-%m-%Y')
    # Add headers
    headers = ['Sr.No.',"Description of Goods","Work start- date","Work end-date","Challan No.","Challan date","call log.no.","call log. Date","RC Code","HSN no.","Location Code","UOM",'Qty.','Rate','Amount',"SGST Rate","SGST Amount",'CGST Rate','CGST Amount','Total Amount payable']
    ws.append(headers)
    for cell in ws["4:4"]:
        cell.font = Font(bold=True)
    # Add data from the model
    dt=InvoiceModel.objects.get(id=id)
    products = InvoiceRCModel.objects.filter(InvoiceNo=dt.InvoiceNo)
    num=1
    for i in products:
        en = EntryModel.objects.get(Counter=i.Counter)
        ws.append([num,i.RCDescription,en.Date.strftime('%d-%m-%Y'),en.CloseDate.strftime('%d-%m-%Y'),en.ChallanNo,en.ChallanDate.strftime('%d-%m-%Y'),en.CodeNo,en.Date.strftime('%d-%m-%Y'),i.RCCode,i.HSNCode,num+1000,i.Unit,i.Quantity,i.Rate,i.Amount,i.GSTRate + " %",i.GSTAmount,i.GSTRate + " %",i.GSTAmount,i.TotalAmount])
        num+=1

    Amount = 0
    GSTAmount = 0
    TotalAmount = 0
    for j in products:
        Amount += eval(j.Amount)
        GSTAmount += eval(j.GSTAmount)
        TotalAmount += eval(j.TotalAmount)
    Amount = "%.2f" % Amount
    GSTAmount = "%.2f" % GSTAmount
    TotalAmount = "%.2f" % TotalAmount
    ws.append(['','','','','','','','','','','','','','',Amount,'',GSTAmount,'',GSTAmount,TotalAmount ])
    side = Side(border_style='thin', color='000000')
    set_border(ws, side)
    # Save the workbook to the HttpResponse
    wb.save(response)
    return response


def HSN(request,id):
    dt=InvoiceModel.objects.get(id=id)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="GA20242{int(dt.InvoiceNo)+5000}_HSN Format.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "HSN_SAC"
    # Add headers
    headers = ['Invoice no',"HSN/SAC CODE",'BASIC AMOUNT','CGST Rate','CGST Amount',"SGST Rate","SGST Amount","IGST Rate","IGST Amount",'TOTAL TAX AMOUNT']
    ws.append(headers)
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    # Add data from the model
    dt=InvoiceModel.objects.get(id=id)
    products = InvoiceRCModel.objects.filter(InvoiceNo=dt.InvoiceNo)
    num=1
    for i in products:
        en = EntryModel.objects.get(Counter=i.Counter)
        IN="GA20242"+str(int(dt.InvoiceNo)+5000)
        ws.append([IN,i.HSNCode,i.Amount,i.GSTRate + " %",i.GSTAmount,i.GSTRate + " %",i.GSTAmount,'0','0',i.TotalAmount])
        num+=1

    Amount = 0
    GSTAmount = 0
    TotalAmount = 0
    for j in products:
        Amount += eval(j.Amount)
        GSTAmount += eval(j.GSTAmount)
        TotalAmount += eval(j.TotalAmount)
    Amount = "%.2f" % Amount
    GSTAmount = "%.2f" % GSTAmount
    TotalAmount = "%.2f" % TotalAmount
    ws.append(['','',Amount,'',GSTAmount,'',GSTAmount,'','',TotalAmount ])

    side = Side(border_style='thin', color='000000')
    set_border(ws, side)
    # Save the workbook to the HttpResponse
    wb.save(response)
    return response

def SoftCopy(request,id):
    dt=InvoiceModel.objects.get(id=id)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="GA20242{int(dt.InvoiceNo)+5000}-Soft Copy.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    headers = ['V.code',"V.Name","Vendor type (MSME/Non-MSME}","Inv.no.","Inv.date.","Nature of Service (R&M, Hospitality, Stationery, Courier, Capex, AMC, 2m3m Hk, 2m3m material, Runnerboys etc.)","Exp. Period","ASD no. in case of R&M","Loc.","Loc. name","CC code","CC name",'Basic Inv. amount.',"SGST Rate","SGST Amount",'CGST Rate','CGST Amount','Total Amount payable']
    ws.append(headers)
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    # Add data from the model
    dt=InvoiceModel.objects.get(id=id)
    products = InvoiceRCModel.objects.filter(InvoiceNo=dt.InvoiceNo)
    num=1
    month = int(dt.InvoiceData.strftime('%m'))-1
    year = int(dt.InvoiceData.strftime('%Y'))-2000
    my = str(month)+'-'+str(year)
    print(my)
    for i in products:
        en = EntryModel.objects.get(Counter=i.Counter)
        ws.append(['278949','Giri Associates','MSME',f'GA20242{int(dt.InvoiceNo)+5000}',dt.InvoiceData.strftime('%d-%m-%Y'),'R & M',my,'ADM0077763','957','Mahuva-Mahuva',en.CostCode,en.CCName,i.Amount,i.GSTRate + " %",i.GSTAmount,i.GSTRate + " %",i.GSTAmount,i.TotalAmount])
        num+=1

    Amount = 0
    GSTAmount = 0
    TotalAmount = 0
    for j in products:
        Amount += eval(j.Amount)
        GSTAmount += eval(j.GSTAmount)
        TotalAmount += eval(j.TotalAmount)
    Amount = "%.2f" % Amount
    GSTAmount = "%.2f" % GSTAmount
    TotalAmount = "%.2f" % TotalAmount
    ws.append(['','','','','','','','','','','','',Amount,'',GSTAmount,'',GSTAmount,TotalAmount ])
    side = Side(border_style='thin', color='000000')
    set_border(ws, side)
    # Save the workbook to the HttpResponse
    wb.save(response)
    return response

@login_required(login_url='Login')
def QuotationP(request,id):
    dt=QuotationModel.objects.get(id=id)
    party = PartyModel.objects.all()
    Rc = QTTModel.objects.filter(Counter=dt.Counter)
    RCAmount = 0
    NAmount = 0
    Amount = 0
    for j in Rc:
        if j.TYPES == 'RCCode':
            RCAmount += eval(j.Amount)
        else:
            NAmount += eval(j.Amount)
        Amount += eval(j.Amount)
    RCAmount = "%.2f" % RCAmount
    NAmount = "%.2f" % NAmount
    Amount = "%.2f" % Amount
    p = {}
    for i in party:
        p['PartyName'] = i.PartyName
        break
    main={'dt':dt,'party':party,'p':p,'Rc':Rc,'RCAmount':RCAmount,'NAmount':NAmount,'Amount':Amount}
    return render(request,'quotationp.html',main)
